"""
counter.py — Module quản lý Log xe Ra/Vào và xuất báo cáo Excel.

Chức năng chính:
- Lưu trữ mảng log xe tạm thời trong bộ nhớ (Python list).
- Thêm bản ghi khi xe cắt vạch (add_log).
- Tính tổng hợp theo loại xe và hướng (get_summary).
- Xuất file Excel với 2 sheet: Summary + Detail (export_to_excel).
"""
from datetime import datetime
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd

# ============================================================
# BẢNG ÁNH XẠ CLASS ID (COCO) → TÊN PHƯƠNG TIỆN
# Chỉ giữ lại 4 lớp phương tiện giao thông cần theo dõi.
# ============================================================
VEHICLE_CLASS_MAP = {
    2: "car",          # Ô tô con
    3: "motorcycle",   # Xe máy
    5: "bus",          # Xe buýt
    7: "truck",        # Xe tải
}


def _frame_to_timestamp(frame_number: int, fps: float) -> str:
    """Chuyển đổi số frame thành chuỗi thời gian HH:MM:SS.

    Args:
        frame_number: Số thứ tự frame trong video.
        fps: Tốc độ khung hình của video (frames per second).

    Returns:
        Chuỗi thời gian định dạng "HH:MM:SS".
    """
    if fps <= 0:
        return "00:00:00"
    total_seconds = int(frame_number / fps)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


class VehicleLogger:
    """Quản lý mảng log xe Ra/Vào tạm thời trong bộ nhớ.

    Mỗi lần một phương tiện cắt qua vạch ảo (LineZone), gọi phương thức
    ``add_log()`` để ghi nhận một bản ghi mới vào mảng ``self._logs``.

    Attributes:
        _logs (list[dict]): Mảng chứa toàn bộ bản ghi log xe.
        _fps (float): Tốc độ khung hình video, dùng để tính timestamp.
        _logged_ids (set[int]): Tập hợp track_id đã ghi log (tránh ghi trùng).
    """

    def __init__(self, fps: float = 30.0):
        """Khởi tạo VehicleLogger.

        Args:
            fps: Tốc độ khung hình của video nguồn (mặc định 30 FPS).
        """
        self._logs: list[dict] = []
        self._fps: float = fps
        self._logged_ids: set[int] = set()

    # ------------------------------------------------------------------ #
    #  THÊM BẢN GHI LOG                                                   #
    # ------------------------------------------------------------------ #
    def add_log(
        self,
        track_id: int,
        class_id: int,
        direction: str,
        frame_number: int,
        confidence: float = 0.0,
    ) -> bool:
        """Thêm một bản ghi xe cắt vạch vào mảng log.

        Mỗi track_id chỉ được ghi **một lần duy nhất** để tránh đếm trùng
        khi xe dao động quanh vạch.

        Args:
            track_id: ID duy nhất do ByteTrack gán.
            class_id: Mã lớp COCO (2, 3, 5, 7).
            direction: Hướng di chuyển — "Entry" hoặc "Exit".
            frame_number: Frame tại thời điểm xe cắt vạch.
            confidence: Độ tin cậy nhận diện YOLO (0.0 → 1.0).

        Returns:
            True nếu ghi thành công, False nếu track_id đã tồn tại
            hoặc class_id không hợp lệ.
        """
        # Kiểm tra track_id đã ghi chưa (chống đếm trùng)
        if track_id in self._logged_ids:
            return False

        # Kiểm tra class_id có trong danh sách phương tiện hợp lệ
        if class_id not in VEHICLE_CLASS_MAP:
            return False

        # Kiểm tra direction hợp lệ
        if direction not in ("Entry", "Exit"):
            return False

        # Tạo bản ghi mới
        log_entry = {
            "track_id": track_id,
            "class_id": class_id,
            "class_name": VEHICLE_CLASS_MAP[class_id],
            "direction": direction,
            "frame_number": frame_number,
            "timestamp": _frame_to_timestamp(frame_number, self._fps),
            "confidence": round(confidence, 2),
        }

        self._logs.append(log_entry)
        self._logged_ids.add(track_id)
        return True

    # ------------------------------------------------------------------ #
    #  TRUY VẤN DỮ LIỆU                                                  #
    # ------------------------------------------------------------------ #
    @property
    def logs(self) -> list[dict]:
        """Trả về toàn bộ mảng log (chỉ đọc)."""
        return self._logs.copy()

    @property
    def total_entry(self) -> int:
        """Tổng số xe đi Vào (Entry)."""
        return sum(1 for log in self._logs if log["direction"] == "Entry")

    @property
    def total_exit(self) -> int:
        """Tổng số xe đi Ra (Exit)."""
        return sum(1 for log in self._logs if log["direction"] == "Exit")

    @property
    def total_count(self) -> int:
        """Tổng số lượt xe qua vạch (Entry + Exit)."""
        return len(self._logs)

    def get_summary(self) -> dict:
        """Tính tổng hợp số xe theo từng loại và hướng.

        Returns:
            Dictionary dạng:
            {
                "car":        {"Entry": 5, "Exit": 3},
                "motorcycle": {"Entry": 2, "Exit": 1},
                "bus":        {"Entry": 1, "Exit": 0},
                "truck":      {"Entry": 0, "Exit": 1},
            }
        """
        summary = {}
        for class_name in VEHICLE_CLASS_MAP.values():
            summary[class_name] = {"Entry": 0, "Exit": 0}

        for log in self._logs:
            name = log["class_name"]
            direction = log["direction"]
            summary[name][direction] += 1

        return summary

    # ------------------------------------------------------------------ #
    #  XUẤT FILE EXCEL                                                     #
    # ------------------------------------------------------------------ #
    def export_to_excel(self, output_path: str) -> str:
        """Xuất toàn bộ log ra file Excel (.xlsx) gồm 2 sheet.

        - Sheet "Summary": Bảng tổng hợp số xe Vào/Ra theo loại.
        - Sheet "Detail":  Danh sách chi tiết từng lượt xe qua vạch.

        Args:
            output_path: Đường dẫn file xlsx đầu ra.
                         Ví dụ: "outputs/report_video_easy.xlsx"

        Returns:
            Đường dẫn tuyệt đối của file đã xuất.
        """
        # Tạo thư mục cha nếu chưa có
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # --- Sheet 1: Summary ---
        summary = self.get_summary()
        summary_rows = []
        total_entry = 0
        total_exit = 0

        for class_name, counts in summary.items():
            entry_count = counts["Entry"]
            exit_count = counts["Exit"]
            summary_rows.append({
                "Loại xe": class_name,
                "Số xe Vào (Entry)": entry_count,
                "Số xe Ra (Exit)": exit_count,
                "Tổng": entry_count + exit_count,
            })
            total_entry += entry_count
            total_exit += exit_count

        # Dòng tổng cộng
        summary_rows.append({
            "Loại xe": "TỔNG CỘNG",
            "Số xe Vào (Entry)": total_entry,
            "Số xe Ra (Exit)": total_exit,
            "Tổng": total_entry + total_exit,
        })

        df_summary = pd.DataFrame(summary_rows)

        # --- Sheet 2: Detail ---
        detail_rows = []
        for idx, log in enumerate(self._logs, start=1):
            detail_rows.append({
                "STT": idx,
                "Track ID": log["track_id"],
                "Loại xe": log["class_name"],
                "Hướng": log["direction"],
                "Frame": log["frame_number"],
                "Thời gian": log["timestamp"],
                "Độ tin cậy": log["confidence"],
            })

        df_detail = pd.DataFrame(detail_rows)

        # Ghi ra file Excel với 2 sheet
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_summary.to_excel(
                writer,
                sheet_name="Summary",
                index=False,
                startrow=2,
            )
            df_detail.to_excel(
                writer,
                sheet_name="Detail",
                index=False,
                startrow=2,
            )

            workbook = writer.book

            header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
            total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(
                left=Side(style="thin", color="FFB7B7B7"),
                right=Side(style="thin", color="FFB7B7B7"),
                top=Side(style="thin", color="FFB7B7B7"),
                bottom=Side(style="thin", color="FFB7B7B7"),
            )

            for sheet_name in ("Summary", "Detail"):
                worksheet = writer.sheets[sheet_name]
                worksheet.sheet_view.showGridLines = False

                # Create title and subtitle rows
                if sheet_name == "Summary":
                    title = "BÁO CÁO TỔNG HỢP XE"
                    subtitle = f"Tạo lúc: {datetime.now():%Y-%m-%d %H:%M:%S}"
                    data_columns = ["Loại xe", "Số xe Vào (Entry)", "Số xe Ra (Exit)", "Tổng"]
                else:
                    title = "BÁO CÁO CHI TIẾT XE"
                    subtitle = f"Tạo lúc: {datetime.now():%Y-%m-%d %H:%M:%S}"
                    data_columns = ["STT", "Track ID", "Loại xe", "Hướng", "Frame", "Thời gian", "Độ tin cậy"]

                last_column = chr(ord("A") + len(data_columns) - 1)
                worksheet.merge_cells(f"A1:{last_column}1")
                worksheet["A1"] = title
                worksheet["A1"].font = Font(size=14, bold=True, color="FFFFFFFF")
                worksheet["A1"].fill = title_fill
                worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

                worksheet.merge_cells(f"A2:{last_column}2")
                worksheet["A2"] = subtitle
                worksheet["A2"].font = Font(size=10, italic=True)
                worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

                header_row = 3
                worksheet.freeze_panes = f"A{header_row + 1}"

                for cell in worksheet[header_row]:
                    cell.font = Font(bold=True, color="FF000000")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Style total row in summary sheet
                if sheet_name == "Summary":
                    total_row = worksheet.max_row
                    for cell in worksheet[total_row]:
                        cell.font = Font(bold=True)
                        cell.fill = total_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border

                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)

                    for cell in column_cells:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))

                    worksheet.column_dimensions[column_letter].width = min(max_length + 4, 32)

                for row in worksheet.iter_rows(min_row=header_row + 1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border

                if sheet_name == "Detail":
                    # Format confidence column as percentage-like value with two decimals
                    confidence_col = data_columns.index("Độ tin cậy") + 1
                    for row in worksheet.iter_rows(min_row=header_row + 1, min_col=confidence_col, max_col=confidence_col):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0.00"

        return os.path.abspath(output_path)

    # ------------------------------------------------------------------ #
    #  XUẤT FILE CSV (BỔ SUNG)                                            #
    # ------------------------------------------------------------------ #
    def export_to_csv(self, output_path: str) -> str:
        """Xuất mảng log ra file CSV (chỉ phần Detail).

        Args:
            output_path: Đường dẫn file csv đầu ra.

        Returns:
            Đường dẫn tuyệt đối của file đã xuất.
        """
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        df = pd.DataFrame(self._logs)
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
        return os.path.abspath(output_path)

    # ------------------------------------------------------------------ #
    #  RESET                                                               #
    # ------------------------------------------------------------------ #
    def reset(self):
        """Xóa toàn bộ log, dùng khi chuyển sang xử lý video mới."""
        self._logs.clear()
        self._logged_ids.clear()

    def __repr__(self) -> str:
        return (
            f"VehicleLogger(total={self.total_count}, "
            f"entry={self.total_entry}, exit={self.total_exit})"
        )
